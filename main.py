import asyncio
import os
from datetime import datetime
from tkinter import Tk, filedialog
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.async_api import async_playwright, Page

# ============================================================
# Environment Configuration
# ============================================================
load_dotenv()

USERNAME       = os.getenv("SYNERION_USERNAME")
PASSWORD       = os.getenv("SYNERION_PASSWORD")
ATTENDANCE_URL = os.getenv("ATTENDANCE_URL")

# ============================================================
# Excel Configuration (0-based column indices)
# ============================================================
COL_DATE  = 0
COL_ENTRY = 3
COL_EXIT  = 4

# ============================================================
# Helpers
# ============================================================
def select_excel_file() -> str:
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    root.destroy()
    if not file_path:
        raise ValueError("No Excel file selected.")
    return file_path


def read_excel_to_list(file_path: str) -> list[tuple[datetime, str, str]]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    workbook  = load_workbook(path, data_only=True)
    worksheet = workbook.active

    records = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row[COL_DATE]:
            break
        try:
            date_str  = str(row[COL_DATE]).strip()
            work_date = datetime.strptime(date_str, "%d/%m/%Y")
        except Exception:
            continue

        entry     = str(row[COL_ENTRY] or "").strip()
        exit_time = str(row[COL_EXIT]  or "").strip()

        # Skip rows with no data at all
        if not entry and not exit_time:
            continue

        if entry:
            records.append((work_date, entry, exit_time))

    return records

async def login(page: Page) -> None:
    print(f"🔐 Logging...")
    await page.goto(ATTENDANCE_URL, wait_until="networkidle")
    
    await page.fill("#UserName", USERNAME)
    await page.fill("#Password", PASSWORD)
    await page.click('input.loginBtn')
    
    await page.wait_for_load_state("networkidle")
    print("✅ Login successful")

async def close_walkme_popup(page: Page):
    """Close the WalkMe tutorial popup if it appears"""
    print("🔍 Checking for WalkMe popup...")
    try:
        # Multiple possible selectors for the close button
        close_button = page.locator('span.walkme-action-destroy-1.wm-close-link').first
        
        if await close_button.is_visible(timeout=5000):
            await close_button.click()
            print("✅ WalkMe popup closed successfully")
            await asyncio.sleep(1.5)
        else:
            print("ℹ️ No WalkMe popup detected")
    except:
        print("ℹ️ WalkMe popup not found or already closed")

async def ensure_correct_month(page: Page, excel_first_date: datetime) -> None:
    try:
        period_text    = await page.locator(".period-number").first.inner_text(timeout=6000)
        current_month  = period_text.strip()
        expected_month = excel_first_date.strftime("%m/%Y")
        print(f"📅 Current period: {current_month} | Excel: {expected_month}")

        if current_month != expected_month:
            print("⚠️ Wrong month! Attempting to change...")
            await page.locator(".btn-next").first.click()
            await asyncio.sleep(2)
    except Exception as e:
        print(f"ℹ️  Could not read current month: {e}")


async def fill_day(page: Page, work_date: datetime, entry_time: str, exit_time: str) -> bool:
    date_str = work_date.strftime("%d/%m")
    print(f"   → Filling {date_str} | {entry_time} → {exit_time}")

    try:
        # Click date cell
        await page.locator(f".text-right.date:has-text('{date_str}')").first.click(timeout=8000)

        # Fill Entry
        if entry_time:
            entry_input = page.locator('sn-time input[type="text"]').nth(0)
            await entry_input.click()
            await entry_input.fill(entry_time)
            await entry_input.press("Tab")

        # Fill Exit
        if exit_time:
            exit_input = page.locator('sn-time input[type="text"]').nth(1)
            await exit_input.click()
            await exit_input.fill(exit_time)
            await exit_input.press("Tab")

        return True

    except Exception as e:
        print(f"   ❌ Failed to fill {date_str}: {e}")
        return False


# ============================================================
# Main
# ============================================================
async def main() -> None:
    # Select Excel before entering async browser context
    print("Please select the Excel file...")
    excel_file = select_excel_file()

    data_list = read_excel_to_list(excel_file)
    if not data_list:
        print("❌ No records found in the Excel file. Aborting.")
        return
    print(f"✅ Loaded {len(data_list)} records from Excel\n")

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=False, args=["--disable-blink-features=AutomationControlled"])
        context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36")
        try:
            page = await context.new_page()

            await login(page)

            # Close WalkMe popup
            await close_walkme_popup(page)

            # Go to the daily attendance page
            print(f"📄 Opening: {ATTENDANCE_URL}")
            await page.goto(ATTENDANCE_URL, wait_until="networkidle")
            await page.wait_for_timeout(4000)

            await ensure_correct_month(page, data_list[0][0])

            ok, fail = 0, 0
            for i, (work_date, entry, exit_t) in enumerate(data_list, 1):
                print(f"[{i}/{len(data_list)}] {work_date.strftime('%d/%m/%Y')}")
                success = await fill_day(page, work_date, entry, exit_t)
                if success:
                    ok += 1
                else:
                    fail += 1

            print(f"\n{'=' * 40}")
            print(f"✅ {ok} succeeded  |  ❌ {fail} failed")
            print(f"{'=' * 40}")
            print("🎉 Done!")

        finally:
            print("\n" + "=" * 60)
            print("IMPORTANT")
            print("Please verify that there are no absences or special hours")
            print("that require manual entry in Synerion, such as:")
            print("- Vacation")
            print("- Sick leave")
            print("- Holiday")
            print("- Child sick leave")
            print("- Military reserve duty")
            print("- Unpaid leave")
            print("- Bereavement leave")
            print("- Special absence hours")
            print("- Overtime categories")
            print("- Any other attendance exception")
            print()
            print("If needed, add them manually before approving")
            print("your attendance report.")
            print("=" * 60 + "\n")

            input("Press Enter to close the browser...")
            await browser.close()


if __name__ == "__main__":
    asyncio.run(main())